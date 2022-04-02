# Processing data 

# Task: Identify the files that are in our folder

import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet import worksheet
import pyexcel as p
import pyexcel_xls as ps
import pyexcel_xlsx as psx
import glob

cwd = os.getcwd()
files = os.listdir(cwd)  # Get all the files in that directory
print("Files in %r: %s" % (cwd, files))

# ------- Task: Modify XLS files to XLSX files 

p.save_book_as(file_name='2009_4.xls', dest_file_name='2009_4.xlsx')
os.remove("2009_4.xls")

# ------- Task: get names and change sheet names

ss=openpyxl.load_workbook("2008_3.xlsx")
print (ss.sheetnames)

# printing the sheet names

ss_sheet = ss['Responsabilidad Civil']
ss_sheet.title = 'Resp. Civil'

ss.save("2010_1.xlsx")

print (ss.sheetnames)

# ------- Task: delete sheets that we don't need

wb=openpyxl.load_workbook("2010_1.xlsx")
print (wb.sheetnames)
sheet_names = wb.sheetnames
list_of_sheets_to_keep = ['Accidentes Personales', 'Agricola', 'Autos', 'Credito a la Vivienda', 'Credito', 'Fen. Hidrometeorologicos','Gastos Medicos', 'Incendio', 'Resp. Civil', 'Salud', 'Terremoto', 'Transporte de Mercancias', 'Vida', 'Pensiones']
for sheetName in wb.sheetnames:
    if sheetName not in list_of_sheets_to_keep:
        del wb[sheetName]
wb.save('2020_1.xlsx')
wb.get_sheet_names()


# -------- Task: Split data and create dataframes by insurance sector

path = os.getcwd()
excel_files = glob.glob(os.path.join(path, "*.xlsx"))
excel_files
  
df = pd.DataFrame()

for f in excel_files:
    try:
        data = pd.read_excel(f, 'Accidentes Personales', skiprows=7)
        df = df.append(data)
    except:
        pass

df.head()


# -------------- Task: clean the data up    

# Delete rows that we don't need.

df = df.drop(df[(df['ENTIDAD'] == "Total") | (df['ENTIDAD'] == "Desconocido") |
(df['ENTIDAD'] == "Desconocido o sin domicil") | (df['ENTIDAD'] == "Desconocido o sin domicilio") |
(df['ENTIDAD'] == "  Extranjero") | (df['ENTIDAD'] == "Extranjero") |
(df['ENTIDAD'] == "No Aplica") | (df['ENTIDAD'] == "Total general")].index)

# Clean whitespaces

df["ENTIDAD"] = df["ENTIDAD"].str.strip()

# Change some names in values

df['ENTIDAD'].replace('Distrito Federal','Ciudad de México',inplace=True)

df["ENTIDAD"].value_counts()  

# Delete Unnamed error columns

df = df[df.columns.drop(list(df.filter(regex='Unnamed:')))]

df.columns

# Replace NaN columns with actual data

df = df.assign(**{
    'RIESGOS ASEGURADOS': df['RIESGOS ASEGURADOS'].fillna(df['ASEGURADOS EN VIGOR'])})

df = df.assign(**{
    'RECLAMACIONES': df['RECLAMACIONES'].fillna(df['SINIESTROS'])})

# Check is there any NaN's values in our worked (filled) columns    

df['RIESGOS ASEGURADOS'].isnull().values.any()
df['RECLAMACIONES'].isnull().values.any()

# Drop used columns for fill

df = df.drop(columns=['ASEGURADOS EN VIGOR', 'SINIESTROS'])

df.columns

# Change the column's name

df = df.rename(columns={'RIESGOS ASEGURADOS': 'RIESGOS_ASEGURADOS'})

# Create column Type for the merge

df["TIPO"] = "Accidentes Personales"

df.shape
df.head()

# ------------ Task: Export to XLSX the split and cleaned dataframes by insurance sector. 

df.to_excel("data_accidentes.xlsx")


# --------------Task: Look for all our data in our folder and Merge all into one dataframes

import os
excel_files = [filename for filename in os.listdir() if filename.startswith("data") and not filename.endswith("data_pensiones.xlsx")]
excel_files

df = pd.DataFrame()

for f in excel_files:
    try:
        data = pd.read_excel(f)
        df = df.append(data)
    except:
        pass

df.columns

# Clean some quick issues

df = df[df.columns.drop(list(df.filter(regex='Unnamed:')))]

# Visualize our clean and worked data, ready for be analyzed.

df.head()

# Export to XLSX the final data.

df.to_excel("data_all.xlsx")


# ---------Task: Load Final DataFrame and plot time series 

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

data = pd.read_excel("data_all.xlsx")

data.head()

# Plotting
fig = plt.figure(figsize = (10, 5))
 
# Creating the bar plot

plt.bar(data["TIPO"], data["RECLAMACIONES"], color ='maroon',
        width = 0.4)
 
plt.xlabel("Tipo de Seguro")
plt.ylabel("Número de Reclamaciones")
plt.title("Numero de Reclamaciones por Tipo de seguro")
plt.show()

# fig.savefig('test.jpg')


# Data ready and Let's Work!

# -------- Task: Add an date column

# df["Date"] = "01012010"
# df['Date'] = pd.to_datetime(df['Date'], format='%d%m%Y')
# df.head()

